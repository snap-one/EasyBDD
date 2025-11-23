"""
Data source loaders for data-driven testing
Supports CSV, JSON, Excel, and inline data
"""

import csv
import json
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
from loguru import logger


class DataLoader:
    """Load test data from various sources"""

    @staticmethod
    def load_from_file(file_path: Path) -> List[Dict[str, Any]]:
        """
        Load data from a file based on its extension

        Args:
            file_path: Path to the data file

        Returns:
            List of dictionaries containing test data

        Raises:
            ValueError: If file format is not supported
            FileNotFoundError: If file doesn't exist
        """
        if not file_path.exists():
            raise FileNotFoundError(f"Data file not found: {file_path}")

        extension = file_path.suffix.lower()

        if extension == ".csv":
            return DataLoader.load_from_csv(file_path)
        elif extension == ".json":
            return DataLoader.load_from_json(file_path)
        elif extension in [".xlsx", ".xls"]:
            return DataLoader.load_from_excel(file_path)
        else:
            raise ValueError(
                f"Unsupported data file format: {extension}. "
                "Supported formats: .csv, .json, .xlsx, .xls"
            )

    @staticmethod
    def load_from_csv(file_path: Path) -> List[Dict[str, Any]]:
        """
        Load data from a CSV file

        Args:
            file_path: Path to CSV file

        Returns:
            List of dictionaries with column headers as keys
        """
        logger.info(f"Loading data from CSV: {file_path}")
        data = []

        with open(file_path, "r", encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                # Convert numeric strings to actual numbers
                processed_row = {}
                for key, value in row.items():
                    processed_row[key] = DataLoader._convert_value(value)
                data.append(processed_row)

        logger.info(f"Loaded {len(data)} rows from CSV")
        return data

    @staticmethod
    def load_from_json(file_path: Path) -> List[Dict[str, Any]]:
        """
        Load data from a JSON file

        Args:
            file_path: Path to JSON file

        Returns:
            List of dictionaries containing test data
        """
        logger.info(f"Loading data from JSON: {file_path}")

        with open(file_path, "r", encoding="utf-8") as jsonfile:
            data = json.load(jsonfile)

        # Handle both list of objects and single object
        if isinstance(data, dict):
            data = [data]
        elif not isinstance(data, list):
            raise ValueError(f"JSON data must be an object or array, got {type(data)}")

        logger.info(f"Loaded {len(data)} items from JSON")
        return data

    @staticmethod
    def load_from_excel(
        file_path: Path, sheet_name: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """
        Load data from an Excel file

        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to load (default: first sheet)

        Returns:
            List of dictionaries with column headers as keys
        """
        logger.info(f"Loading data from Excel: {file_path}")

        workbook = openpyxl.load_workbook(file_path, data_only=True)

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet_name}' not found. "
                    f"Available sheets: {', '.join(workbook.sheetnames)}"
                )
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active

        data = []
        headers = None

        for i, row in enumerate(worksheet.iter_rows(values_only=True)):
            if i == 0:
                # First row is headers
                headers = [
                    str(cell) if cell is not None else f"Column{j}"
                    for j, cell in enumerate(row)
                ]
                continue

            if all(cell is None for cell in row):
                # Skip empty rows
                continue

            row_data = {}
            for header, cell_value in zip(headers, row):
                row_data[header] = DataLoader._convert_value(cell_value)

            data.append(row_data)

        workbook.close()
        logger.info(f"Loaded {len(data)} rows from Excel")
        return data

    @staticmethod
    def _convert_value(value: Any) -> Any:
        """
        Convert string values to appropriate types

        Args:
            value: Value to convert

        Returns:
            Converted value (int, float, bool, or string)
        """
        if not isinstance(value, str):
            return value

        value = value.strip()

        # Check for boolean
        if value.lower() in ["true", "yes", "1"]:
            return True
        if value.lower() in ["false", "no", "0"]:
            return False

        # Check for null/none
        if value.lower() in ["null", "none", ""]:
            return None

        # Try to convert to int
        try:
            if "." not in value:
                return int(value)
        except ValueError:
            pass

        # Try to convert to float
        try:
            return float(value)
        except ValueError:
            pass

        # Return as string
        return value


class DataGenerator:
    """Generate test data combinations"""

    @staticmethod
    def combine_data(data_sources: List[List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
        """
        Combine multiple data sources into cartesian product

        Args:
            data_sources: List of data source lists

        Returns:
            List of combined dictionaries
        """
        if not data_sources:
            return []

        if len(data_sources) == 1:
            return data_sources[0]

        # Cartesian product of all data sources
        from itertools import product

        combined = []
        for combination in product(*data_sources):
            merged_dict = {}
            for data_dict in combination:
                merged_dict.update(data_dict)
            combined.append(merged_dict)

        return combined

    @staticmethod
    def filter_data(
        data: List[Dict[str, Any]], condition: str, variables: Dict[str, Any] = None
    ) -> List[Dict[str, Any]]:
        """
        Filter data based on a condition expression

        Args:
            data: List of data dictionaries
            condition: Python expression to evaluate
            variables: Additional variables for evaluation context

        Returns:
            Filtered list of dictionaries
        """
        from ..core.safe_eval import safe_eval

        filtered = []
        for item in data:
            context = {**(variables or {}), **item}
            try:
                if safe_eval(condition, context):
                    filtered.append(item)
            except Exception as e:
                logger.warning(f"Failed to evaluate condition for item {item}: {e}")

        return filtered
